#!/usr/bin/env python

#=========================================================================================
# Author: Patrick Brockmann CEA/DRF/LSCE - October 2024
#=========================================================================================

import sys
import os 
import re 
import numpy as np
import pandas as pd

from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QIcon

import matplotlib.pyplot as plt
from matplotlib.patches import ConnectionPatch
from matplotlib.lines import Line2D
from matplotlib.axis import XAxis, YAxis
import matplotlib.patches as patches
from matplotlib.colors import is_color_like

from scipy import interpolate

from openpyxl.utils import get_column_letter

#=========================================================================================
if len(sys.argv[1:]) >= 1:
    fileData = sys.argv[1]
else:
    fileData = None

#========================================================================================
version = 'v2.6'
serie1Color = 'darkmagenta'
serie2Color = 'forestgreen'
pointerColor = 'blue'
curveWidth = 0.8

#========================================================================================
detached_windows = {}
tableDataWidget = None
tablePointersWidget = None

#========================================================================================
fig = None 
axs = None
axsInterp = None
curve1 = None
curve2 = None
points1 = None
points2 = None
curve2Interp = None
vline1 = None
vline2 = None
dfData = None
x1 = []
y1 = [] 
x2 = []
y2 = [] 
x2Interp = [] 
showInterp = False
linecursor1 = None
linecursor2 = None
x1Name = None
y1Name = None 
x2Name = None
y2Name = None
artistsList_Dict = {} 
vline1List = []
vline2List = []
coordsX1 = []
coordsX2 = []
coordsX1Name = None
coordsX2Name = None
second_xaxis = None
figPointers = None 
axPointers = None 
axGradient = None 
y1_axisInverted = False
y2_axisInverted = False

press = None
cur_xlim = None
cur_ylim = None
xpress = None
ypress = None
mousepress = None
press_origin = None
key_x = None
key_shift = None
key_control = None
artistLabel = None

combo_x1 = None
combo_y1 = None
combo_x2 = None
combo_y2 = None
button_c1 = None
button_c2 = None
check_y1 = None
check_y2 = None
buttonGroup_i3 = None

f_1to2 = None
f_2to1 = None
interpMode = 'linear'

#========================================================================================
def create_Data_tab():
    global tableDataWidget

    widget = QWidget()
    layout = QVBoxLayout()

    tableDataWidget = QTableWidget()
    tableDataWidget.setEditTriggers(QTableWidget.NoEditTriggers)

    layout.addWidget(tableDataWidget)
    widget.setLayout(layout)
    return widget

#========================================================================================
def create_Settings_tab():
    global combo_x1, combo_y1, combo_x2, combo_y2, button_c1, button_c2, check_y1, check_y2
    global buttonGroup_i3

    widget = QWidget()
    layout = QVBoxLayout()

    #----------------------------------------
    groupbox1 = QGroupBox('Serie #1')
    groupbox1_layout = QVBoxLayout()

    layout_x1 = QHBoxLayout()
    label_x1 = QLabel('X1 :')
    combo_x1 = QComboBox()
    combo_x1.setFixedSize(400, 30)
    layout_x1.addWidget(label_x1)
    layout_x1.addWidget(combo_x1)
    layout_x1.addStretch()

    layout_y1 = QHBoxLayout()
    label_y1 = QLabel('Y1 :')
    combo_y1 = QComboBox()
    combo_y1.setFixedSize(400, 30)
    check_y1 = QCheckBox('y axis inverted')
    check_y1.clicked.connect(reverseAxis)
    layout_y1.addWidget(label_y1)
    layout_y1.addWidget(combo_y1)
    layout_y1.addWidget(check_y1)
    layout_y1.addStretch()

    layout_c1 = QHBoxLayout()
    label_c1 = QLabel('Color :')
    button_c1 = QPushButton()
    button_c1.setStyleSheet(f'background-color: {serie1Color};')
    button_c1.clicked.connect(lambda: changeColor(1))
    layout_c1.addWidget(label_c1)
    layout_c1.addWidget(button_c1)
    layout_c1.addStretch()

    groupbox1_layout.addLayout(layout_x1)
    groupbox1_layout.addLayout(layout_y1)
    groupbox1_layout.addLayout(layout_c1)

    groupbox1.setLayout(groupbox1_layout)

    #----------------------------------------
    groupbox2 = QGroupBox('Serie #2')
    groupbox2_layout = QVBoxLayout()

    layout_x2 = QHBoxLayout()
    label_x2 = QLabel('X2 :')
    combo_x2 = QComboBox()
    combo_x2.setFixedSize(400, 30)
    layout_x2.addWidget(label_x2)
    layout_x2.addWidget(combo_x2)
    layout_x2.addStretch()

    layout_y2 = QHBoxLayout()
    label_y2 = QLabel('Y2 :')
    combo_y2 = QComboBox()
    combo_y2.setFixedSize(400, 30)
    check_y2 = QCheckBox('y axis inverted')
    check_y2.clicked.connect(reverseAxis)
    layout_y2.addWidget(label_y2)
    layout_y2.addWidget(combo_y2)
    layout_y2.addWidget(check_y2)
    layout_y2.addStretch()

    layout_c2 = QHBoxLayout()
    label_c2 = QLabel('Color :')
    button_c2 = QPushButton()
    button_c2.setStyleSheet(f'background-color: {serie2Color};')
    button_c2.clicked.connect(lambda: changeColor(2))
    layout_c2.addWidget(label_c2)
    layout_c2.addWidget(button_c2)
    layout_c2.addStretch()

    groupbox2_layout.addLayout(layout_x2)
    groupbox2_layout.addLayout(layout_y2)
    groupbox2_layout.addLayout(layout_c2)

    groupbox2.setLayout(groupbox2_layout)

    #----------------------------------------
    groupbox3 = QGroupBox('Interpolation')
    groupbox3_layout = QVBoxLayout()

    buttonGroup_i3 = QButtonGroup()
    buttonGroup_i3.buttonClicked.connect(changeInterpMode)

    linear_radio = QRadioButton("Linear interpolation")
    #linear_radio.setToolTip("")
    linear_radio.setChecked(True)
    buttonGroup_i3.addButton(linear_radio, 0)
    groupbox3_layout.addWidget(linear_radio)

    pchip_radio = QRadioButton("Piecewise Cubic Hermite Interpolating Polynomial (PCHIP) interpolation")
    #pchip_radio.setToolTip("")
    buttonGroup_i3.addButton(pchip_radio, 1)
    groupbox3_layout.addWidget(pchip_radio)
    
    groupbox3.setLayout(groupbox3_layout)

    #----------------------------------------
    layout.addWidget(groupbox1)
    layout.addWidget(groupbox2)
    layout.addWidget(groupbox3)
    layout.addStretch()

    #----------------------------------------
    with open('style.css', 'r') as f:
            css = f.read()
            widget.setStyleSheet(css)

    widget.setLayout(layout)
    return widget

#========================================================================================
def create_Pointers_tab():
    global tablePointersWidget
    global figPointers, axPointers

    widget = QWidget()
    layout = QHBoxLayout()

    #---------------------------------------------------------
    column_width = 200
    tablePointersWidget = QTableWidget()
    tablePointersWidget.setRowCount(0)
    tablePointersWidget.setHorizontalHeaderLabels([coordsX1Name, coordsX2Name])
    tablePointersWidget.setColumnCount(2)
    tablePointersWidget.setColumnWidth(0, column_width)
    tablePointersWidget.setColumnWidth(1, column_width)
    tablePointersWidget.verticalHeader().setFixedWidth(50)
    tablePointersWidget.horizontalHeader().setDefaultAlignment(Qt.AlignRight)
    tablePointersWidget.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
    tablePointersWidget.setEditTriggers(QTableWidget.NoEditTriggers)

    total_width = (column_width* 2) + tablePointersWidget.verticalHeader().width() + tablePointersWidget.frameWidth() * 2
    tablePointersWidget.setFixedWidth(total_width)

    #---------------------------------------------------------
    figPointers, axPointers = plt.subplots(1, 1)
    figPointers.subplots_adjust(left=0.15, right=0.85, top=0.85, bottom=0.15)
    figPointers.set_visible(False)

    figPointers.canvas.setFocusPolicy(Qt.ClickFocus)

    #---------------------------------------------------------
    layout.addWidget(tablePointersWidget)
    layout.addWidget(figPointers.canvas)

    widget.setLayout(layout)
    return widget

#========================================================================================
def create_Plots_tab():
    global fig, axs 

    widget = QWidget()
    layout = QVBoxLayout()

    #---------------------------------------------------------
    fig, axs = plt.subplots(2, 1)
    fig.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.1, wspace=0, hspace=0.2)
    fig.set_visible(False)

    fig.canvas.setFocusPolicy(Qt.ClickFocus)

    fig.canvas.mpl_connect('button_press_event', on_mouse_press)
    fig.canvas.mpl_connect('button_release_event', on_mouse_release)
    fig.canvas.mpl_connect('scroll_event', on_mouse_scroll)
    fig.canvas.mpl_connect('motion_notify_event', on_mouse_motion)
    fig.canvas.mpl_connect('pick_event', on_mouse_pick)
    fig.canvas.mpl_connect('key_press_event', on_key_press)
    fig.canvas.mpl_connect('key_release_event', on_key_release)

    #---------------------------------------------------------
    layout.addWidget(fig.canvas)

    widget.setLayout(layout)
    return widget

#=========================================================================================
def updatePlots():
    global fig, axs, axsInterp
    global curve1, curve2, points1, points2, linecursor1, linecursor2

    activate_tab_by_name(tabs, 'Plots')
    fig.set_visible(True)
    fig.canvas.setFocus()

    axs[0].clear()
    axs[1].clear()
    if axsInterp: fig.delaxes(axsInterp)

    #---------------------------------------------------------
    curve1, = axs[0].plot(x1, y1, color=serie1Color, picker=True, pickradius=20, linewidth=curveWidth, label='curve1')
    points1 = axs[0].scatter(x1, y1, s=5, marker='o', color=serie1Color, picker=True, pickradius=20, label='points1')
    points1.set_visible(False)
    linecursor1 = axs[0].axvline(color='k', alpha=0.25, linewidth=1)
    axs[0].grid(visible=True, which='major', color='lightgray', linestyle='dashed', linewidth=0.5)
    axs[0].set_xlabel(x1Name, color=serie1Color)
    axs[0].set_ylabel(y1Name, color=serie1Color)
    axs[0].patch.set_alpha(0)
    axs[0].yaxis.set_inverted(y1_axisInverted)
    axs[0].autoscale()
    axs[0].set_label('curve1')
    axs[0].xaxis.pickradius = 50
    axs[0].yaxis.pickradius = 50
    
    #---------------------------------------------------------
    curve2, = axs[1].plot(x2, y2, color=serie2Color, picker=True, pickradius=20, linewidth=curveWidth, label='curve2')
    points2 = axs[1].scatter(x2, y2, s=5, marker='o', color=serie2Color, picker=True, pickradius=20, label='points2')
    points2.set_visible(False)
    linecursor2 = axs[1].axvline(color='k', alpha=0.25, linewidth=1)
    axs[1].grid(visible=True, which='major', color='lightgray', linestyle='dashed', linewidth=0.5)
    axs[1].set_xlabel(x2Name, color=serie2Color)
    axs[1].set_ylabel(y2Name, color=serie2Color)
    axs[1].yaxis.set_inverted(y2_axisInverted)
    axs[1].autoscale()
    axs[1].set_label('curve2')
    axs[1].xaxis.pickradius = 50
    axs[1].yaxis.pickradius = 50
    
    #---------------------------------------------------------
    axsInterp = axs[0].twinx()
    axsInterp.sharey(axs[1])
    axsInterp.set_ylabel(y2Name, color=serie2Color)
    axsInterp.set_zorder(-10)
    axsInterp.set_visible(showInterp)
    axsInterp.set_label('curve2Interp')

    #---------------------------------------------------------
    #print('axs0', axs[0].get_xlim())
    #print('axs1', axs[1].get_xlim())

#=========================================================================================
def openData():
    fileName, _ = QFileDialog.getOpenFileName(main_window, 'Open Excel file', '', 'Excel files (*.xlsx);;All files (*)')
    if not fileName: 
        displayStatusMessage('Data', 'Cannot open file', 5000)
        return
    #print('-------------------', fileName)
    loadData(fileName)

#=========================================================================================
def getVariables():
    global x1Name, y1Name, x2Name, y2Name
    global x1, y1, x2, y2
    global coordsX1Name, coordsX2Name

    x1Name = combo_x1.currentText()
    y1Name = combo_y1.currentText()
    x2Name = combo_x2.currentText()
    y2Name = combo_y2.currentText()
    x1 = dfData[x1Name].to_numpy()
    y1 = dfData[y1Name].to_numpy()
    x2 = dfData[x2Name].to_numpy()
    y2 = dfData[y2Name].to_numpy()

    coordsX1Name = x1Name
    coordsX2Name = x2Name

#=========================================================================================
def changeVariables():

    deleteInterp()
    deleteConnections()

    getVariables()
    updatePlots()
    drawConnections()
    setInterp()
    displayInterp(showInterp)
    updateAxes()

#=========================================================================================
def loadData(fileName):
    global coordsX1, coordsX2, coordsX1Name, coordsX2Name
    global tableDataWidget
    global showInterp 
    global combo_x1, combo_y1, combo_x2, combo_y2
    global dfData
    global serie1Color, serie2Color, y1_axisInverted, y2_axisInverted

    activate_tab_by_name(tabs, 'Data')
    #--------------------------------------------
    try:
        dfData = pd.read_excel(fileName)
    except:
        displayStatusMessage('Main', 'Cannot open file', 5000)
        return

    if len(dfData.columns) < 4:
        displayStatusMessage('Main', 'File does not contain 4 columns', 5000)
        return

    #--------------------------------------------
    combo_x1.addItems(dfData.columns)  
    combo_y1.addItems(dfData.columns)  
    combo_x2.addItems(dfData.columns)  
    combo_y2.addItems(dfData.columns) 
    
    #--------------------------------------------
    try:
        readingMessages = ''
        dfSettings = pd.read_excel(fileName, sheet_name='Settings')
        read_x1Name = dfSettings['Coordinates X'][0]
        read_y1Name = dfSettings['Coordinates Y'][0]
        read_x2Name = dfSettings['Coordinates X'][1]
        read_y2Name = dfSettings['Coordinates Y'][1]
        if ((read_x1Name not in dfData.columns) or
            (read_y1Name not in dfData.columns) or
            (read_x2Name not in dfData.columns) or
            (read_y2Name not in dfData.columns)):
            combo_x1.setCurrentIndex(0)         # Take 4 first variables 
            combo_y1.setCurrentIndex(1)
            combo_x2.setCurrentIndex(2)
            combo_y2.setCurrentIndex(3)
            readingMessages += 'In Settings sheet, cannot find variables, set 4 first' + '<br>'
        else:
            x1Name = read_x1Name
            y1Name = read_y1Name
            x2Name = read_x2Name
            y2Name = read_y2Name

        read_serie1Color = dfSettings['Colors'][0]
        read_serie2Color = dfSettings['Colors'][1]
        if not is_color_like(read_serie1Color):
            readingMessages += 'In Settings sheet, unknown item in "Colors": ' + read_serie1Color + '<br>'
        else:
            serie1Color = read_serie1Color
        if not is_color_like(read_serie2Color):
            readingMessages += 'In Settings sheet, unknown item in "Colors": ' + read_serie2Color + '<br>'
        else:
            serie2Color = read_serie2Color

        read_y1_axisInverted = dfSettings['Y axis inverted'][0]
        if isinstance(read_y1_axisInverted, (bool, np.bool)):
            y1_axisInverted = bool(read_y1_axisInverted)
        else:
            readingMessages += 'In Settings sheet, unknown item in "Y axis inverted": ' + str(read_y1_axisInverted) + '<br>'

        read_y2_axisInverted = dfSettings['Y axis inverted'][1]
        if isinstance(read_y2_axisInverted, (bool, np.bool)):
            y2_axisInverted = bool(read_y2_axisInverted)
        else:
            readingMessages += 'In Settings sheet, unknown item in "Y axis inverted": ' + str(read_y2_axisInverted) + '<br>'

        combo_x1.setCurrentText(x1Name)
        combo_y1.setCurrentText(y1Name)
        combo_x2.setCurrentText(x2Name)
        combo_y2.setCurrentText(y2Name)
        check_y1.setChecked(y1_axisInverted)
        check_y2.setChecked(y2_axisInverted)
        button_c1.setStyleSheet(f'background-color: {serie1Color};')
        button_c2.setStyleSheet(f'background-color: {serie2Color};')

    except:
        readingMessages += 'Sheetname "Settings" not found or misformatted' + '<br>'
        combo_x1.setCurrentIndex(0)         # Take 4 first variables 
        combo_y1.setCurrentIndex(1)
        combo_x2.setCurrentIndex(2)
        combo_y2.setCurrentIndex(3)

    #--------------------------------------------
    if readingMessages != '':
        QMessageBox.warning(
            main_window, 
            'Messages from reading', 
            readingMessages,
            QMessageBox.Ok, 
            QMessageBox.Ok
        )

    #--------------------------------------------
    combo_x1.currentIndexChanged.connect(changeVariables)
    combo_y1.currentIndexChanged.connect(changeVariables)
    combo_x2.currentIndexChanged.connect(changeVariables)
    combo_y2.currentIndexChanged.connect(changeVariables)

    #--------------------------------------------
    deleteConnections()
    deleteCoords()
    updatePointers()
    deleteInterp()
    showInterp = False

    getVariables()
    updatePlots()
    updatePointers()
    updateAxes()

    tableDataWidget.clearContents()
    tableDataWidget.setRowCount(dfData.shape[0])
    tableDataWidget.setColumnCount(dfData.shape[1])
    tableDataWidget.setHorizontalHeaderLabels(dfData.columns)
    tableDataWidget.verticalHeader().setFixedWidth(50)
    tableDataWidget.resizeColumnsToContents()

    for row in range(dfData.shape[0]):
        for col in range(dfData.shape[1]):
            value = str(dfData.iat[row, col])
            tableDataWidget.setItem(row, col, QTableWidgetItem(value))
            item = tableDataWidget.item(row, col)
            if row % 2 == 0:
                item.setBackground(QColor(250, 250, 250))

    #--------------------------------------------
    try:
        dfPointers = pd.read_excel(fileName, sheet_name='Pointers')
        coordsX1Name = dfPointers.columns[0]
        coordsX2Name = dfPointers.columns[1]
        coordsX1 = dfPointers[coordsX1Name].to_numpy()
        coordsX2 = dfPointers[coordsX2Name].to_numpy()

       # check if arrays are monotonically increasing
        if not (((np.diff(coordsX1) >= 0).all()) and ((np.diff(coordsX2) >= 0).all())):
            displayStatusMessage('Data', 'Error: pointer coordinates are not monotonically increasing', 5000)
            coordsX1 = []
            coordsX2 = []
        else:
            drawConnections()
            updatePointers()
            updateAxes()

    except:
        displayStatusMessage('Data', 'No sheetname Pointers found', 5000)

#=========================================================================================
def updateConnections():

    for artistsList in artistsList_Dict.values():
        if isinstance(artistsList[0], ConnectionPatch):
            connect = artistsList[0]
            x1, y1 = connect.xy1
            connect.xy1 = (x1, axs[0].get_ylim()[0])
            x2, y2 = connect.xy2
            connect.xy2 = (x2, axs[1].get_ylim()[1])
            if ((axs[0].get_xlim()[0] < x1 < axs[0].get_xlim()[1]) and
                (axs[1].get_xlim()[0] < x2 < axs[1].get_xlim()[1])):
                connect.set_visible(True)
            else:
                connect.set_visible(False)

#=========================================================================================
def drawConnections():
    global artistsList_Dict, vline1List, vline2List

    for i in range(len(coordsX1)):
        coordX1 = coordsX1[i]
        coordX2 = coordsX2[i]
        vline1 = axs[0].axvline(coordX1, color=pointerColor, alpha=0.5, linestyle='--', linewidth=1, label='vline1')
        vline2 = axs[1].axvline(coordX2, color=pointerColor, alpha=0.5, linestyle='--', linewidth=1, label='vline2')
        vline1List.append(vline1)
        vline2List.append(vline2)
        connect = ConnectionPatch(color=pointerColor, alpha=0.5, linewidth=1, picker=10, clip_on=True, label='connection',
                    xyA=(coordX1, axs[0].get_ylim()[0]), coordsA=axs[0].transData,
                    xyB=(coordX2, axs[1].get_ylim()[1]), coordsB=axs[1].transData)
        fig.add_artist(connect)
        artistsList_Dict[id(connect)] = [connect, vline1, vline2]

    updateConnections()

#=========================================================================================
def deleteConnections():
    global artistsList_Dict, vline1List, vline2List

    for objectId in artistsList_Dict.keys():
        for artist in artistsList_Dict[objectId]:
            artist.remove()
    artistsList_Dict = {}
    vline1List = []
    vline2List = []

#=========================================================================================
def deleteCoords():
    global coordsX1, coordsX2
    
    coordsX1 = []
    coordsX2 = []

#=========================================================================================
def updatePointers():
    global coordsX1, coordsX2
    global tablePointersWidget
    global figPointers, axPointers, axGradient

    coordsX1 = sorted([float(line.get_xdata()[0]) for line in vline1List])
    coordsX2 = sorted([float(line.get_xdata()[0]) for line in vline2List])

    tablePointersWidget.clearContents()
    tablePointersWidget.setHorizontalHeaderLabels([coordsX1Name, coordsX2Name])
    tablePointersWidget.setRowCount(len(coordsX1))

    for row in range(len(coordsX1)):
        tablePointersWidget.setItem(row, 0, QTableWidgetItem(f'{coordsX1[row]:.4f}'))
        tablePointersWidget.setItem(row, 1, QTableWidgetItem(f'{coordsX2[row]:.4f}'))
        item1 = tablePointersWidget.item(row, 0)
        item1.setTextAlignment(Qt.AlignRight)
        item2 = tablePointersWidget.item(row, 1)
        item2.setTextAlignment(Qt.AlignRight)
        if row % 2 == 0: 
            item1.setBackground(QColor(250, 250, 250))
            item2.setBackground(QColor(250, 250, 250))

    if len(coordsX1) >= 2:
        axPointers.clear()
        if axGradient: figPointers.delaxes(axGradient)
        figPointers.set_visible(True)

        axPointers.plot(coordsX1, coordsX2, color='steelblue', linewidth=1)
        axPointers.scatter(coordsX1, coordsX2, s=10, marker='o', color='steelblue')
        axPointers.grid(visible=True, which='major', color='lightgray', linestyle='dashed', linewidth=1)
        axPointers.set_xlabel(x1Name, color=serie1Color)
        axPointers.set_ylabel(x2Name, color=serie2Color)
        axPointers.autoscale()

        defineInterpFunction()
        gradient = np.gradient(f_1to2(x1), x1).astype(np.float32)       # to avoid unnecessary precision
        axGradient = axPointers.twinx()
        axGradient.set_ylabel('Gradient', color='darkorange')
        axGradient.plot(x1, gradient, color='darkorange', lw=1)
        axGradient.autoscale(axis='y', tight=False)

        figPointers.canvas.draw()
    else:
        axPointers.clear()
        if axGradient: 
            figPointers.delaxes(axGradient)
            axGradient = None
        figPointers.set_visible(False)
        figPointers.canvas.draw()

    displayStatusMessage('Plots', 'Pointers updated', 5000)

#=========================================================================================
def updateAxes():
    # Full vertical range on both plots, horizontal range set from pointers 
    
    # autoscale to get ylim range
    axs[0].relim()
    axs[1].relim()
    axs[0].autoscale(axis='y')
    axs[1].autoscale(axis='y')
    ylim_axs0 = axs[0].get_ylim()
    ylim_axs1 = axs[1].get_ylim()
    
    # hide 
    displayInterp(False)
    linecursor1.set_visible(False)
    linecursor2.set_visible(False)

    # autoscale to get xlim range only from pointers
    curve1.set_visible(False)
    curve2.set_visible(False)
    axs[0].relim(visible_only=True)
    axs[1].relim(visible_only=True)
    axsInterp.relim(visible_only=True)
    axs[0].autoscale()
    axs[1].autoscale()
    axsInterp.autoscale()
    xlim_axs0 = axs[0].get_xlim()
    xlim_axs1 = axs[1].get_xlim()
    
    # apply xlim and ylim ranges
    curve1.set_visible(True)
    curve2.set_visible(True)
    axs[0].set_xlim(xlim_axs0)
    axs[1].set_xlim(xlim_axs1)
    axs[0].set_ylim(ylim_axs0)
    axs[1].set_ylim(ylim_axs1)
    
    updateConnections()
    displayInterp(showInterp)
    fig.canvas.draw()

#========================================================================================
def on_key_press(event):
    global key_x, key_shift, key_control, vline1, vline2, artistsList_Dict
    global tablePointersWidget
    global showInterp

    sys.stdout.flush()

    #-----------------------------------------------
    if event.key == 'a':
        updateAxes()

    #-----------------------------------------------
    if event.key == 'x':
        key_x = True

    #-----------------------------------------------
    if event.key == 'X':
        reply = QMessageBox.question(
            main_window, 
            'Confirmation', 
            'Are you sure you want to delete all pointers ?', 
            QMessageBox.Yes | QMessageBox.No, 
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            deleteConnections()
            deleteCoords()
            updatePointers()
            deleteInterp()
            showInterp = False
            displayInterp(showInterp)
            displayStatusMessage('Plots', 'Pointers deleted', 5000)

    #-----------------------------------------------
    elif event.key == 'z':
        showInterp = not showInterp
        setInterp()
        displayInterp(showInterp)

    #-----------------------------------------------
    elif event.key == 'shift':
        key_shift = True

    #-----------------------------------------------
    elif event.key == 'control':
        key_control = True
        if event.inaxes == axs[0]:
            points1.set_visible(True)
        elif event.inaxes == axs[1]:
            points2.set_visible(True)
        fig.canvas.draw()

    #-----------------------------------------------
    if event.key == 'c':
        if vline1 != None and vline2 != None :
            coordX1 = float(vline1.get_xdata()[0])
            coordX2 = float(vline2.get_xdata()[0])
            # current coordsX1, coordsX2. Will be defined later from setInterp
            coordsX1_cur = sorted([float(line.get_xdata()[0]) for line in vline1List])
            coordsX2_cur = sorted([float(line.get_xdata()[0]) for line in vline2List])
            # Check positions
            if np.searchsorted(coordsX1_cur, coordX1) != np.searchsorted(coordsX2_cur, coordX2):
                #print('Error: Connection not possible because it would cross existing connections')
                displayStatusMessage('Plots', 'Error: Connection not possible because it would cross existing connections', 5000)
                return

            connect = ConnectionPatch(color=pointerColor, alpha=0.5, linewidth=1, picker=10, clip_on=True, label='connection',
                        xyA=(coordX1, axs[0].get_ylim()[0]), coordsA=axs[0].transData,
                        xyB=(coordX2, axs[1].get_ylim()[1]), coordsB=axs[1].transData)
            fig.add_artist(connect)
            artistsList_Dict[id(connect)] = [connect, vline1, vline2]
            vline1List.append(vline1)
            vline2List.append(vline2)
            vline1 = None
            vline2 = None

            updatePointers()
            fig.canvas.draw()

            if len(vline1List) >= 2:
                setInterp()
                displayInterp(showInterp)

#------------------------------------------------------------------
def on_key_release(event):
    global key_x, key_shift, key_control

    sys.stdout.flush()

    #-----------------------------------------------
    if event.key == 'x':
        key_x = False

    #-----------------------------------------------
    elif event.key == 'shift':
        key_shift = False

    #-----------------------------------------------
    elif event.key == 'control':
        key_control = False
        if event.inaxes == axs[0]:
            points1.set_visible(False)
        elif event.inaxes == axs[1]:
            points2.set_visible(False)
        fig.canvas.draw()

#------------------------------------------------------------------
def on_mouse_pick(event):
    global artistLabel
    global vline1, vline2, artistsList_Dict, vline1List, vline2List

    artistLabel = event.artist.get_label()

    #-----------------------------------------------
    if artistLabel == 'connection':
        if key_x:
            objectId = id(event.artist)
            for artist in artistsList_Dict[objectId]:
                artist.remove()
                if artist in vline1List:
                    vline1List.remove(artist)
                if artist in vline2List:
                    vline2List.remove(artist)
            del artistsList_Dict[objectId]
            
            updatePointers()
            setInterp()
            displayInterp(showInterp)

    #-----------------------------------------------
    if artistLabel in ['curve1', 'curve2']:
        if key_shift:
            coordPoint = [event.mouseevent.xdata, event.mouseevent.ydata]
            if event.artist == curve1:
                if vline1 != None:
                    vline1.set_data([coordPoint[0], coordPoint[0]], [0,1])
                else:
                    vline1 = axs[0].axvline(coordPoint[0], color=pointerColor, alpha=0.5, linestyle='--', linewidth=1, label='vline1')
            elif event.artist == curve2:
                if vline2 != None:
                    vline2.set_data([coordPoint[0], coordPoint[0]], [0,1])
                else:
                    vline2 = axs[1].axvline(coordPoint[0], color=pointerColor, alpha=0.5, linestyle='--', linewidth=1, label='vline2')
            fig.canvas.draw()

    #-----------------------------------------------
    elif artistLabel in ['points1', 'points2']:
        if key_control:
            ind = event.ind[0]
            if event.artist == points1:
                coordPoint = [x1[ind], y1[ind]]
                if vline1 != None:
                    vline1.set_data([coordPoint[0], coordPoint[0]], [0,1])
                else:
                    vline1 = axs[0].axvline(coordPoint[0], color=pointerColor, linestyle='--', linewidth=1, label='vline1')
            elif event.artist == points2:
                coordPoint = [x2[ind], y2[ind]]
                if vline2 != None:
                    vline2.set_data([coordPoint[0], coordPoint[0]], [0,1])
                else:
                    vline2 = axs[1].axvline(coordPoint[0], color=pointerColor, linestyle='--', linewidth=1, label='vline2')
            fig.canvas.draw()

#------------------------------------------------------------------
def on_mouse_press(event):
    global cur_xlim, cur_ylim, press, xpress, ypress, mousepress, press_origin
    global serie1Color, serie2Color
    global curve1, curve2, curve2Interp

    if event.inaxes not in axs: return

    mousepress = None
    if event.button == 1:
        mousepress = 'left'
        press_origin = event.inaxes

    cur_xlim = event.inaxes.get_xlim()
    cur_ylim = event.inaxes.get_ylim()
    press = event.xdata, event.ydata
    xpress, ypress = press

#------------------------------------------------------------------
def on_mouse_release(event):
    global press
    press = None

#------------------------------------------------------------------
def on_mouse_motion(event):
    global cur_xlim, cur_ylim, press

    #-----------------------------------------------
    if not fig.get_visible():
        return 

    #-----------------------------------------------
    found = False
    for artistsList in artistsList_Dict.values():
        if isinstance(artistsList[0], ConnectionPatch):
            connect = artistsList[0]
            if connect.contains(event)[0]:
                connect.set_color('red')
                found = True 
            else:
                connect.set_color(pointerColor)
    if found:
        fig.canvas.draw_idle()
        return

    #-----------------------------------------------
    if event.inaxes not in axs:
        press = None
        linecursor1.set_visible(False)
        linecursor2.set_visible(False)
        points1.set_visible(False)
        points2.set_visible(False)
        fig.canvas.draw_idle()
        return

    #-----------------------------------------------
    if event.inaxes is axs[0]:
        linecursor1.set_visible(True)
        linecursor2.set_visible(False)
        points2.set_visible(False)
        linecursor1.set_xdata([event.xdata])
    elif event.inaxes is axs[1]:
        linecursor1.set_visible(False)
        points1.set_visible(False)
        linecursor2.set_visible(True)
        linecursor2.set_xdata([event.xdata])
    fig.canvas.draw_idle()

    #-----------------------------------------------
    if press is None: return

    #-----------------------------------------------
    # When mousepress has been done not in the listen axe
    if event.inaxes.get_label() != press_origin.get_label(): return

    #-----------------------------------------------
    if mousepress == 'left':
        linecursor1.set_visible(False)
        linecursor2.set_visible(False)
        dx = event.xdata - xpress
        dy = event.ydata - ypress
        cur_xlim -= dx
        cur_ylim -= dy
        event.inaxes.set_xlim(cur_xlim[0], cur_xlim[1])
        event.inaxes.set_ylim(cur_ylim[0], cur_ylim[1])

        updateConnections()
        event.inaxes.figure.canvas.draw()
        event.inaxes.figure.canvas.flush_events()

#------------------------------------------------------------------
def detect_artist(event):
    """Detect which artist the scroll event occurred on, including axis ticks."""
    for ax in axs:
        # First check if the event occurred on the X or Y axis ticks
        if ax.xaxis.contains(event)[0]:
            return ax.xaxis  # Detected XAxis
        if ax.yaxis.contains(event)[0]:
            return ax.yaxis  # Detected YAxis

        # Now check if the event occurred inside the Axes itself
        if ax.contains(event)[0]:
            return ax  # Detected Axes itself
    return None

#------------------------------------------------------------------
def on_mouse_scroll(event):

    """Callback for scroll event to detect artist on the correct axis."""
    artist = detect_artist(event)  # Detect the artist in the entire figure

    if artist is None:
        #print('No artist detected under the scroll event.')
        return

    scale_factor = 0.9 if event.button == 'up' else 1.1

    if isinstance(artist, XAxis):
        ax = artist.axes
        cur_xlim = ax.get_xlim()
        xdata = (cur_xlim[0] + cur_xlim[1]) / 2
        new_xlim = [xdata - (xdata - cur_xlim[0]) * scale_factor,
                    xdata + (cur_xlim[1] - xdata) * scale_factor]
        ax.set_xlim(new_xlim)
        #print(f'Zoomed X axis in axis "{ax.get_label()}"')
        #displayStatusMessage('Plots', 'Zoom on Xaxis', 5000)

    elif isinstance(artist, YAxis):
        ax = artist.axes
        cur_ylim = ax.get_ylim()
        ydata = (cur_ylim[0] + cur_ylim[1]) / 2
        new_ylim = [ydata - (ydata - cur_ylim[0]) * scale_factor,
                    ydata + (cur_ylim[1] - ydata) * scale_factor]
        ax.set_ylim(new_ylim)
        #print(f'Zoomed Y axis in axis "{ax.get_label()}"')
        #displayStatusMessage('Plots', 'Zoom on Yaxis', 5000)

    elif isinstance(artist, plt.Axes):
        ax = artist
        cur_xlim = ax.get_xlim()
        cur_ylim = ax.get_ylim()
        xdata = event.xdata
        ydata = event.ydata

        # Only zoom in the plot area, hence the need to check xdata and ydata
        #if xdata is None or ydata is None:
        #    return  # Prevent NoneType error if data coords are unavailable

        new_xlim = [xdata - (xdata - cur_xlim[0]) * scale_factor,
                    xdata + (cur_xlim[1] - xdata) * scale_factor]
        new_ylim = [ydata - (ydata - cur_ylim[0]) * scale_factor,
                    ydata + (cur_ylim[1] - ydata) * scale_factor]
        ax.set_xlim(new_xlim)
        ax.set_ylim(new_ylim)
        #print(f'Zoomed both axes in axis "{ax.get_label()}"')
        #displayStatusMessage('Plots', 'Zoom on axes', 5000)

    updateConnections()
    ax.figure.canvas.draw()

#=========================================================================================
def changeColor(serieNumber):
    global serie1Color, serie2Color

    color = QColorDialog.getColor()
    if color.isValid():
        if serieNumber == 1:
            serie1Color = color.name()
            button_c1.setStyleSheet(f'background-color: {serie1Color};')
            curve1.set_color(serie1Color)
            points1.set_color(serie1Color)
            axs[0].set_xlabel(x1Name, c=serie1Color)
            axs[0].set_ylabel(y1Name, c=serie1Color)
            axPointers.set_xlabel(x1Name, color=serie1Color)
        elif serieNumber == 2:
            serie2Color = color.name()
            button_c2.setStyleSheet(f'background-color: {serie2Color};')
            curve2.set_color(serie2Color)
            points2.set_color(serie2Color)
            if curve2Interp: curve2Interp.set_color(serie2Color)
            axs[1].set_xlabel(x2Name, color=serie2Color)
            axs[1].set_ylabel(y2Name, color=serie2Color)
            axsInterp.set_ylabel(y2Name, color=serie2Color)
            second_xaxis.set_xlabel(x2Name, color=serie2Color)
            axPointers.set_ylabel(x2Name, color=serie2Color)
        figPointers.canvas.draw()
        fig.canvas.draw()
        fig.canvas.flush_events()

#=========================================================================================
def reverseAxis():
    global y1_axisInverted, y2_axisInverted

    fig.canvas.setFocus()
    y1_axisInverted = check_y1.isChecked()
    y2_axisInverted = check_y2.isChecked()
    axs[0].yaxis.set_inverted(y1_axisInverted)
    axs[1].yaxis.set_inverted(y2_axisInverted)
    updateConnections()
    fig.canvas.draw()

#=========================================================================================
def deleteInterp():
    global x2Interp, curve2Interp, second_xaxis, showInterp
    global f_1to2, f_2to1

    if curve2Interp:
        curve2Interp.remove()
        curve2Interp = None
        second_xaxis.remove()
        second_xaxis = None
        x2Interp = []
        f_1to2 = None
        f_2to1 = None

#=========================================================================================
def changeInterpMode(button):
    global interpMode

    interpMode_id = buttonGroup_i3.id(button)
    if interpMode_id == 0:
        interpMode = 'linear'
    elif interpMode_id == 1:
        interpMode = 'PCHIP'

    defineInterpFunction()
    changeVariables()

#=========================================================================================
def defineInterpFunction():
    global f_1to2, f_2to1

    if interpMode == 'linear':
        f_1to2 = interpolate.interp1d(coordsX1, coordsX2, kind='linear', fill_value='extrapolate')
        f_2to1 = interpolate.interp1d(coordsX2, coordsX1, kind='linear', fill_value='extrapolate')
    elif interpMode == 'PCHIP':
        f_1to2 = interpolate.PchipInterpolator(coordsX1, coordsX2, extrapolate=True)
        f_2to1 = interpolate.PchipInterpolator(coordsX2, coordsX1, extrapolate=True)

#=========================================================================================
def setInterp():
    global x2Interp, curve2Interp, second_xaxis, coordsX1, coordsX2, showInterp

    if len(vline1List) <= 1:
        displayStatusMessage('Plots', 'Warning: interpolation needs a minimum of 2 connections', 5000)
        if showInterp:
            showInterp = False
            displayInterp(showInterp)
        return

    deleteInterp()
    coordsX1 = sorted([float(line.get_xdata()[0]) for line in vline1List])
    coordsX2 = sorted([float(line.get_xdata()[0]) for line in vline2List])
    defineInterpFunction()

    second_xaxis = axsInterp.secondary_xaxis('top', functions=(f_1to2, f_2to1))
    second_xaxis.tick_params(labelrotation=30)
    second_xaxis.set_xlabel(x2Name, color=serie2Color)
    plt.setp(second_xaxis.get_xticklabels(), horizontalalignment='left')

    x2Interp = f_2to1(x2)
    curve2Interp, = axsInterp.plot(x2Interp, y2, color=serie2Color, alpha=0.8, linewidth=curveWidth, label='curve2Interp')

#=========================================================================================
def displayInterp(visible):

    if visible:
            if curve2Interp: curve2Interp.set_visible(True)
            if axsInterp: axsInterp.set_visible(True)
    else:
            if curve2Interp: curve2Interp.set_visible(False)
            if axsInterp: axsInterp.set_visible(False)
    fig.canvas.draw()

#========================================================================================
def saveData():
    fileName, _ = QFileDialog.getSaveFileName(main_window, 'Save Data', '', 'Excel files (*.xlsx)')

    if fileName:

        if not fileName.endswith(".xlsx"):
            fileName += ".xlsx"

        with pd.ExcelWriter(fileName) as writer:
            #---------------------------------------
            dfData.to_excel(writer, sheet_name='Data', index=False, float_format='%.8f')
            worksheet = writer.sheets['Data']
            for i, col in enumerate(dfData.columns, 1): 
                worksheet.column_dimensions[get_column_letter(i)].width = 25 

            #---------------------------------------
            df = pd.DataFrame({'Coordinates X': [x1Name, x2Name], 
                               'Coordinates Y': [y1Name, y2Name],
                               'Colors': [serie1Color, serie2Color],
                               'Y axis inverted': [y1_axisInverted, y2_axisInverted]
                               })
            df.to_excel(writer, sheet_name='Settings', index=False)
            worksheet = writer.sheets['Settings']
            for i, col in enumerate(df.columns, 1): 
                worksheet.column_dimensions[get_column_letter(i)].width = 25 

            #---------------------------------------
            df = pd.DataFrame({coordsX1Name: coordsX1, coordsX2Name: coordsX2})
            df.to_excel(writer, sheet_name='Pointers', index=False, float_format='%.8f')
            worksheet = writer.sheets['Pointers']
            for i, col in enumerate(df.columns, 1): 
                worksheet.column_dimensions[get_column_letter(i)].width = 25 

            #---------------------------------------
            if len(x2Interp) > 0:
                df = pd.DataFrame({x1Name: x1, y1Name: y1, x2Name: x2, y2Name: y2,
                        y2Name + ' interpolated (' + interpMode + ') on ' + x1Name: x2Interp})
                df.to_excel(writer, sheet_name='Interpolation', index=False, float_format='%.8f')
                worksheet = writer.sheets['Interpolation']
                for i, col in enumerate(df.columns, 1): 
                    worksheet.column_dimensions[get_column_letter(i)].width = 25 

        displayStatusMessage('Data', 'Saved Data in file ' + fileName, 5000)

#========================================================================================
def savePlots():
    fileName, _ = QFileDialog.getSaveFileName(main_window, 'Save Plots', '', 'PNG Files (*.png);;PDF Files (*.pdf)')
    if fileName:
        plt.savefig(fileName)
        displayStatusMessage('Plots', 'Saved Plots in file ' + fileName, 5000)

#========================================================================================
def detach_tab(tab_widget, index):
    tab_content = tab_widget.widget(index)
    tab_name = tab_widget.tabText(index)

    tab_widget.removeTab(index)

    detached_window = QMainWindow()
    detached_window.setWindowTitle(tab_name)
    detached_window.setGeometry(200, 200, 1200, 800)

    detached_window.setCentralWidget(tab_content)
    detached_window.statusBar().showMessage(f'{tab_name} - Ready', 3000)

    tab_content.show()

    def on_close_event(event):
        reattached_content = detached_window.takeCentralWidget()
        tab_widget.addTab(reattached_content, tab_name)
        tab_widget.setCurrentWidget(reattached_content)
        del detached_windows[tab_name]

    detached_window.closeEvent = on_close_event

    detached_window.show()
    detached_windows[tab_name] = detached_window

#========================================================================================
def activate_tab_by_name(tab_widget, tab_name):
    for index in range(tab_widget.count()):
        if tab_widget.tabText(index) == tab_name:
            tab_widget.setCurrentIndex(index)
            return

#========================================================================================
def show_dialog(title, fileHTML, width, height):
    with open(fileHTML, 'r') as file:
        help_text = file.read()
    
    dialog = QDialog()
    dialog.setWindowTitle(title)
    dialog.setFixedSize(width, height)
    
    main_layout = QVBoxLayout()
    dialog.setLayout(main_layout)
    text_browser = QTextBrowser()
    text_browser.setHtml(help_text)
    main_layout.addWidget(text_browser)
    
    button_layout = QHBoxLayout()
    button_layout.addItem(QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
    ok_button = QPushButton('OK')
    ok_button.clicked.connect(dialog.accept)
    icon = QApplication.style().standardIcon(QStyle.SP_DialogApplyButton)
    ok_button.setIcon(icon)
    button_layout.addWidget(ok_button)
    main_layout.addLayout(button_layout)

    dialog.exec_()

#========================================================================================
def displayStatusMessage(target, message, duration=0):

    if target in detached_windows:
        detached_windows[target].statusBar().showMessage(message, duration)
    else:
        main_window.statusBar().showMessage(message, duration)

#========================================================================================
app = QApplication(sys.argv)

icon = QIcon('PyAnalyseries_icon.png')
app.setWindowIcon(icon)

main_window = QMainWindow()
main_window.setGeometry(200, 200, 1200, 800)

tabs = QTabWidget(main_window)
tabs.setMovable(True)

tab1 = create_Data_tab()
tab2 = create_Settings_tab()
tab3 = create_Plots_tab()
tab4 = create_Pointers_tab()
tabs.addTab(tab1, 'Data')
tabs.addTab(tab2, 'Settings')
tabs.addTab(tab3, 'Plots')
tabs.addTab(tab4, 'Pointers')

tabs.tabBarDoubleClicked.connect(lambda index: detach_tab(tabs, index))
#tabs.currentChanged.connect(lambda index: displayStatusMessage('Main', tabs.tabText(index), 2000))

main_window.setCentralWidget(tabs)
menu_bar = main_window.menuBar()

file_menu = menu_bar.addMenu('File')
help_menu = menu_bar.addMenu('Help')
about_menu = menu_bar.addMenu('About')

openData_action = QAction('Open Data', main_window)
saveData_action = QAction('Save Data', main_window)
savePlots_action = QAction('Save Plots', main_window)
exit_action = QAction('Exit', main_window)
exit_action.setShortcut('Q') 
openData_action.triggered.connect(openData)
saveData_action.triggered.connect(saveData)
savePlots_action.triggered.connect(savePlots)
exit_action.triggered.connect(app.quit)
file_menu.addAction(openData_action)
file_menu.addSeparator()
file_menu.addAction(saveData_action)
file_menu.addAction(savePlots_action)
file_menu.addSeparator()
file_menu.addAction(exit_action)

help_action = QAction('Help', main_window)
help_action.triggered.connect(lambda: show_dialog('Help', 'help.html', 800, 700))
help_menu.addAction(help_action)

about_action = QAction('About', main_window)
about_action.triggered.connect(lambda: show_dialog('About', 'about.html', 800, 600))
about_menu.addAction(about_action)

if fileData:
    #print('-------------------', fileData)
    displayStatusMessage('Main', fileData + ' loading', 5000)
    loadData(fileData)

main_window.setStatusBar(QStatusBar())
displayStatusMessage('Main', 'Application ready', 5000)
main_window.show()

sys.exit(app.exec_())
